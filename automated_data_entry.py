import os
import re
import json
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd

# ==================== AUTOMATED DATA ENTRY - BUSINESS GOLD MINE ====================

class AutomatedDataEntry:
    def __init__(self):
        self.ocr_engine = None
        self.extraction_templates = self._load_extraction_templates()
        self.processing_history = []
        
    def _load_extraction_templates(self):
        """Load data extraction templates for different document types"""
        return {
            'receipt': {
                'patterns': {
                    'total': [r'total[:\s]*\$?(\d+\.?\d*)', r'amount[:\s]*\$?(\d+\.?\d*)'],
                    'date': [r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', r'(\d{4}-\d{2}-\d{2})'],
                    'merchant': [r'^([A-Z\s&]+)(?=\n|\r)', r'store[:\s]*([A-Za-z\s]+)'],
                    'items': [r'(\d+\.?\d*)\s*x?\s*([A-Za-z\s]+)\s*\$?(\d+\.?\d*)']
                },
                'excel_headers': ['Date', 'Merchant', 'Item', 'Quantity', 'Price', 'Total']
            },
            'invoice': {
                'patterns': {
                    'invoice_number': [r'invoice[#\s]*:?\s*(\w+)', r'inv[#\s]*:?\s*(\w+)'],
                    'date': [r'date[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})'],
                    'total': [r'total[:\s]*\$?(\d+\.?\d*)', r'amount\s+due[:\s]*\$?(\d+\.?\d*)'],
                    'vendor': [r'from[:\s]*([A-Za-z\s&]+)', r'bill\s+to[:\s]*([A-Za-z\s&]+)'],
                    'line_items': [r'(\d+\.?\d*)\s*([A-Za-z\s]+)\s*\$?(\d+\.?\d*)']
                },
                'excel_headers': ['Invoice #', 'Date', 'Vendor', 'Description', 'Amount', 'Total']
            },
            'business_card': {
                'patterns': {
                    'name': [r'^([A-Z][a-z]+\s+[A-Z][a-z]+)', r'([A-Z][a-z]+\s+[A-Z]\.\s+[A-Z][a-z]+)'],
                    'email': [r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'],
                    'phone': [r'(\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4})', r'(\+\d{1,3}[-.\s]?\d{3,4}[-.\s]?\d{3,4}[-.\s]?\d{3,4})'],
                    'company': [r'([A-Z][A-Za-z\s&]+(?=\n|\r))', r'company[:\s]*([A-Za-z\s&]+)'],
                    'title': [r'([A-Z][a-z]+\s+[A-Z][a-z]+)', r'title[:\s]*([A-Za-z\s]+)']
                },
                'excel_headers': ['Name', 'Title', 'Company', 'Email', 'Phone', 'Address']
            },
            'table': {
                'patterns': {
                    'headers': [r'^([A-Za-z\s]+)(?=\t|\s{2,})', r'([A-Z][A-Za-z\s]+)(?=\|)'],
                    'rows': [r'([^\t\n]+)(?=\t|$)', r'([^|\n]+)(?=\||$)']
                },
                'excel_headers': ['Column1', 'Column2', 'Column3', 'Column4', 'Column5']
            }
        }
    
    def process_receipt_image(self, image_path, output_excel=None):
        """Take receipt/invoice image → Extract to Excel"""
        try:
            if not os.path.exists(image_path):
                return f"Image file not found: {image_path}"
            
            # Extract text from image using OCR
            extracted_text = self._extract_text_from_image(image_path)
            
            if not extracted_text:
                return "Could not extract text from image. Please ensure image is clear and readable."
            
            # Parse receipt data
            receipt_data = self._parse_receipt_data(extracted_text)
            
            # Create Excel file
            if not output_excel:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_excel = os.path.join(os.path.expanduser("~"), "Desktop", f"receipt_data_{timestamp}.xlsx")
            
            self._create_receipt_excel(receipt_data, output_excel)
            
            # Save to processing history
            self._save_processing_record('receipt', image_path, output_excel, receipt_data)
            
            return f"✅ Receipt processed successfully!\n\nExtracted Data:\n{self._format_receipt_summary(receipt_data)}\n\nExcel file saved: {output_excel}"
        
        except Exception as e:
            return f"Error processing receipt: {str(e)}"
    
    def process_pdf_table(self, pdf_path, output_excel=None):
        """Read PDF table → Convert to spreadsheet"""
        try:
            if not os.path.exists(pdf_path):
                return f"PDF file not found: {pdf_path}"
            
            # Extract text from PDF
            extracted_text = self._extract_text_from_pdf(pdf_path)
            
            if not extracted_text:
                return "Could not extract text from PDF."
            
            # Parse table data
            table_data = self._parse_table_data(extracted_text)
            
            # Create Excel file
            if not output_excel:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                pdf_name = os.path.splitext(os.path.basename(pdf_path))[0]
                output_excel = os.path.join(os.path.expanduser("~"), "Desktop", f"{pdf_name}_table_{timestamp}.xlsx")
            
            self._create_table_excel(table_data, output_excel)
            
            # Save to processing history
            self._save_processing_record('pdf_table', pdf_path, output_excel, table_data)
            
            return f"✅ PDF table processed successfully!\n\nExtracted {len(table_data)} rows\n\nExcel file saved: {output_excel}"
        
        except Exception as e:
            return f"Error processing PDF table: {str(e)}"
    
    def process_business_cards_batch(self, images_folder, output_excel=None):
        """Take 50 business cards → Create contacts spreadsheet"""
        try:
            if not os.path.exists(images_folder):
                return f"Folder not found: {images_folder}"
            
            # Find all image files
            image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.gif']
            image_files = []
            
            for ext in image_extensions:
                image_files.extend(Path(images_folder).glob(f"*{ext}"))
                image_files.extend(Path(images_folder).glob(f"*{ext.upper()}"))
            
            if not image_files:
                return f"No image files found in {images_folder}"
            
            contacts_data = []
            processed_count = 0
            
            for image_file in image_files[:50]:  # Limit to 50 as mentioned
                try:
                    # Extract text from business card
                    extracted_text = self._extract_text_from_image(str(image_file))
                    
                    if extracted_text:
                        # Parse business card data
                        card_data = self._parse_business_card_data(extracted_text)
                        card_data['source_file'] = image_file.name
                        contacts_data.append(card_data)
                        processed_count += 1
                
                except Exception as e:
                    print(f"Error processing {image_file}: {e}")
                    continue
            
            if not contacts_data:
                return "No business card data could be extracted from the images."
            
            # Create Excel file
            if not output_excel:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_excel = os.path.join(os.path.expanduser("~"), "Desktop", f"business_contacts_{timestamp}.xlsx")
            
            self._create_contacts_excel(contacts_data, output_excel)
            
            # Save to processing history
            self._save_processing_record('business_cards', images_folder, output_excel, contacts_data)
            
            return f"✅ Business cards processed successfully!\n\nProcessed: {processed_count} cards\nExtracted: {len(contacts_data)} contacts\n\nExcel file saved: {output_excel}"
        
        except Exception as e:
            return f"Error processing business cards: {str(e)}"
    
    def _extract_text_from_image(self, image_path):
        """Extract text from image using OCR"""
        try:
            # Try to use Tesseract OCR
            try:
                import pytesseract
                from PIL import Image
                
                # Open and preprocess image
                image = Image.open(image_path)
                
                # Convert to RGB if necessary
                if image.mode != 'RGB':
                    image = image.convert('RGB')
                
                # Extract text
                extracted_text = pytesseract.image_to_string(image)
                return extracted_text.strip()
            
            except ImportError:
                # Fallback: Try to use easyocr
                try:
                    import easyocr
                    reader = easyocr.Reader(['en'])
                    results = reader.readtext(image_path)
                    
                    # Combine all detected text
                    extracted_text = ' '.join([result[1] for result in results])
                    return extracted_text.strip()
                
                except ImportError:
                    return "OCR library not available. Please install pytesseract or easyocr."
        
        except Exception as e:
            return f"Error extracting text from image: {str(e)}"
    
    def _extract_text_from_pdf(self, pdf_path):
        """Extract text from PDF"""
        try:
            # Try PyPDF2 first
            try:
                import PyPDF2
                with open(pdf_path, 'rb') as file:
                    reader = PyPDF2.PdfReader(file)
                    text = ""
                    for page in reader.pages:
                        text += page.extract_text() + "\n"
                    return text.strip()
            
            except ImportError:
                # Try pdfplumber as alternative
                try:
                    import pdfplumber
                    with pdfplumber.open(pdf_path) as pdf:
                        text = ""
                        for page in pdf.pages:
                            text += page.extract_text() + "\n"
                        return text.strip()
                
                except ImportError:
                    return "PDF processing library not available. Please install PyPDF2 or pdfplumber."
        
        except Exception as e:
            return f"Error extracting text from PDF: {str(e)}"
    
    def _parse_receipt_data(self, text):
        """Parse receipt data from extracted text"""
        try:
            template = self.extraction_templates['receipt']
            data = {
                'merchant': 'Unknown',
                'date': 'Unknown',
                'total': '0.00',
                'items': []
            }
            
            # Extract merchant
            for pattern in template['patterns']['merchant']:
                match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
                if match:
                    data['merchant'] = match.group(1).strip()
                    break
            
            # Extract date
            for pattern in template['patterns']['date']:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    data['date'] = match.group(1).strip()
                    break
            
            # Extract total
            for pattern in template['patterns']['total']:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    data['total'] = match.group(1).strip()
                    break
            
            # Extract items (simplified)
            lines = text.split('\n')
            for line in lines:
                # Look for lines with prices
                price_match = re.search(r'\$?(\d+\.?\d*)', line)
                if price_match and len(line.strip()) > 5:
                    item_name = re.sub(r'\$?\d+\.?\d*', '', line).strip()
                    if item_name:
                        data['items'].append({
                            'name': item_name,
                            'price': price_match.group(1)
                        })
            
            return data
        
        except Exception as e:
            return {'error': str(e)}
    
    def _parse_business_card_data(self, text):
        """Parse business card data from extracted text"""
        try:
            template = self.extraction_templates['business_card']
            data = {
                'name': 'Unknown',
                'title': 'Unknown',
                'company': 'Unknown',
                'email': 'Unknown',
                'phone': 'Unknown',
                'address': 'Unknown'
            }
            
            # Extract email
            for pattern in template['patterns']['email']:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    data['email'] = match.group(1).strip()
                    break
            
            # Extract phone
            for pattern in template['patterns']['phone']:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    data['phone'] = match.group(1).strip()
                    break
            
            # Extract name (first line that looks like a name)
            lines = text.split('\n')
            for line in lines:
                line = line.strip()
                if len(line) > 3 and len(line) < 50:
                    # Check if it looks like a name (2-3 words, proper case)
                    words = line.split()
                    if 2 <= len(words) <= 3 and all(word[0].isupper() for word in words if word):
                        data['name'] = line
                        break
            
            # Extract company (look for company-like words)
            for line in lines:
                line = line.strip()
                if any(word in line.lower() for word in ['inc', 'llc', 'corp', 'company', 'ltd', '&']):
                    data['company'] = line
                    break
            
            return data
        
        except Exception as e:
            return {'error': str(e)}
    
    def _parse_table_data(self, text):
        """Parse table data from extracted text"""
        try:
            lines = text.split('\n')
            table_data = []
            
            # Simple table parsing - look for lines with multiple columns
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Try different delimiters
                columns = None
                for delimiter in ['\t', '|', '  ', ',']:
                    if delimiter in line:
                        columns = [col.strip() for col in line.split(delimiter) if col.strip()]
                        if len(columns) >= 2:
                            break
                
                if columns and len(columns) >= 2:
                    table_data.append(columns)
            
            return table_data
        
        except Exception as e:
            return []
    
    def _create_receipt_excel(self, data, output_path):
        """Create Excel file for receipt data"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Receipt Data"
            
            # Headers
            headers = ['Date', 'Merchant', 'Item', 'Price', 'Total']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data
            row = 2
            if 'items' in data and data['items']:
                for item in data['items']:
                    ws.cell(row=row, column=1, value=data.get('date', 'Unknown'))
                    ws.cell(row=row, column=2, value=data.get('merchant', 'Unknown'))
                    ws.cell(row=row, column=3, value=item.get('name', 'Unknown'))
                    ws.cell(row=row, column=4, value=item.get('price', '0.00'))
                    ws.cell(row=row, column=5, value=data.get('total', '0.00') if row == 2 else '')
                    row += 1
            else:
                # Single row if no items detected
                ws.cell(row=row, column=1, value=data.get('date', 'Unknown'))
                ws.cell(row=row, column=2, value=data.get('merchant', 'Unknown'))
                ws.cell(row=row, column=3, value='Total Purchase')
                ws.cell(row=row, column=4, value=data.get('total', '0.00'))
                ws.cell(row=row, column=5, value=data.get('total', '0.00'))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
        
        except Exception as e:
            raise Exception(f"Error creating receipt Excel: {str(e)}")
    
    def _create_contacts_excel(self, contacts_data, output_path):
        """Create Excel file for business contacts"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Business Contacts"
            
            # Headers
            headers = ['Name', 'Title', 'Company', 'Email', 'Phone', 'Source File']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data
            for row, contact in enumerate(contacts_data, 2):
                ws.cell(row=row, column=1, value=contact.get('name', 'Unknown'))
                ws.cell(row=row, column=2, value=contact.get('title', 'Unknown'))
                ws.cell(row=row, column=3, value=contact.get('company', 'Unknown'))
                ws.cell(row=row, column=4, value=contact.get('email', 'Unknown'))
                ws.cell(row=row, column=5, value=contact.get('phone', 'Unknown'))
                ws.cell(row=row, column=6, value=contact.get('source_file', 'Unknown'))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
        
        except Exception as e:
            raise Exception(f"Error creating contacts Excel: {str(e)}")
    
    def _create_table_excel(self, table_data, output_path):
        """Create Excel file for table data"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Extracted Table"
            
            # Write data
            for row_idx, row_data in enumerate(table_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    
                    # Make first row bold (headers)
                    if row_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
        
        except Exception as e:
            raise Exception(f"Error creating table Excel: {str(e)}")
    
    def _format_receipt_summary(self, data):
        """Format receipt data for display"""
        try:
            summary = f"Merchant: {data.get('merchant', 'Unknown')}\n"
            summary += f"Date: {data.get('date', 'Unknown')}\n"
            summary += f"Total: ${data.get('total', '0.00')}\n"
            
            if 'items' in data and data['items']:
                summary += f"Items: {len(data['items'])} detected\n"
                for i, item in enumerate(data['items'][:3], 1):  # Show first 3 items
                    summary += f"  {i}. {item.get('name', 'Unknown')} - ${item.get('price', '0.00')}\n"
                
                if len(data['items']) > 3:
                    summary += f"  ... and {len(data['items']) - 3} more items\n"
            
            return summary
        
        except Exception as e:
            return f"Error formatting summary: {str(e)}"
    
    def _save_processing_record(self, process_type, input_path, output_path, extracted_data):
        """Save processing record to history"""
        try:
            record = {
                'type': process_type,
                'input_path': input_path,
                'output_path': output_path,
                'timestamp': datetime.now().isoformat(),
                'data_summary': self._create_data_summary(process_type, extracted_data)
            }
            
            self.processing_history.append(record)
        
        except Exception as e:
            print(f"Error saving processing record: {e}")
    
    def _create_data_summary(self, process_type, data):
        """Create summary of extracted data"""
        try:
            if process_type == 'receipt':
                return f"Merchant: {data.get('merchant', 'Unknown')}, Items: {len(data.get('items', []))}"
            elif process_type == 'business_cards':
                return f"Contacts extracted: {len(data)}"
            elif process_type == 'pdf_table':
                return f"Rows extracted: {len(data)}"
            else:
                return "Data processed"
        
        except Exception as e:
            return "Summary unavailable"
    
    def get_processing_history(self):
        """Get processing history"""
        try:
            if not self.processing_history:
                return "No data processing history found"
            
            result = "📊 Data Processing History:\n\n"
            for i, record in enumerate(self.processing_history[-10:], 1):  # Show last 10
                timestamp = datetime.fromisoformat(record['timestamp']).strftime('%Y-%m-%d %H:%M')
                result += f"{i}. {record['type'].upper()}\n"
                result += f"   Input: {os.path.basename(record['input_path'])}\n"
                result += f"   Output: {os.path.basename(record['output_path'])}\n"
                result += f"   Summary: {record['data_summary']}\n"
                result += f"   Processed: {timestamp}\n\n"
            
            return result
        
        except Exception as e:
            return f"Error getting processing history: {str(e)}"

# ==================== GLOBAL INSTANCE ====================

data_entry_engine = AutomatedDataEntry()

# ==================== CONVENIENCE FUNCTIONS ====================

def process_receipt_to_excel(image_path, output_excel=None):
    """Process receipt image to Excel"""
    return data_entry_engine.process_receipt_image(image_path, output_excel)

def process_pdf_table_to_excel(pdf_path, output_excel=None):
    """Process PDF table to Excel"""
    return data_entry_engine.process_pdf_table(pdf_path, output_excel)

def process_business_cards_to_excel(images_folder, output_excel=None):
    """Process business cards to Excel"""
    return data_entry_engine.process_business_cards_batch(images_folder, output_excel)

def get_data_processing_history():
    """Get data processing history"""
    return data_entry_engine.get_processing_history()
